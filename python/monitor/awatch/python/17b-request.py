import asyncio, websockets, openpyxl, json, time
from openpyxl import load_workbook
from watchfiles import awatch

# Excel to Web, Class Example
class Xl2WebExample:
  def __init__(self, filepath, filename, site, port):
    # save initial parameter
    self.filepath = filepath
    self.filename = filename
    self.site = site
    self.port = port

  # miscellanous helper
  def __pack_data(self, fullname):
    # reopen the worksheet all over again
    wb = load_workbook(self.filename)
    ws = wb["Example"]

    return {
      "time": time.ctime(),
      "file": fullname,
      "val1": ws['C2'].value,
      "val2": ws['C3'].value,
      "val3": ws['C4'].value
     }

  def __dump_data(self, data):
    print("Timestamp     : %s" % data["time"])
    print("File Modified : %s" % data["file"])
    print("Data 1: %s" % data["val1"])
    print("Data 2: %s" % data["val2"])
    print("Data 3: %s" % data["val3"])
    print()

  # websocket related
  async def __send_data(self, fullname):
    event_data = self.__pack_data(fullname)
    self.__dump_data(event_data)
    await self.websocket.send(
      json.dumps(event_data))

  async def __monitor_localfile(self):
    async for changes in awatch(self.filepath):
      xlsx = self.filepath + '/' + self.filename

      for change in changes:
        if change[1] == xlsx:
          print(change[0])
          await self.__send_data(change[1])

  async def __monitor_webclient(self):
    while True:
      message = await self.websocket.recv()
      print(message)
      await self.__send_data(None)

  # websocket handler
  async def __handler(self, websocket, path):
    self.websocket = websocket

    task_localfile = asyncio.create_task(
      self.__monitor_localfile())

    task_webclient = asyncio.create_task(
      self.__monitor_webclient())

    # run these two coroutines concurrently
    await(task_localfile)
    await(task_webclient)

  async def main(self):
    # Start the server
    server = await websockets.serve(
      self.__handler, self.site, self.port)
    await server.wait_closed()

# Program Entry Point
example = Xl2WebExample(
  '/home/epsi/awatch/code', 'test-b.xlsx',
  'localhost', 8767)
asyncio.run(example.main())
