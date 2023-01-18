import asyncio, openpyxl, time
import xl2web_25

from openpyxl    import load_workbook
from rich.table  import Table

# Excel to Web, Example Class
class Xl2Web_d(xl2web_25.Xl2Web):
  # Excel Reader
  def _pack_data(self, fullname):
    # reopen the worksheet all over again
    wb = load_workbook(self.xlsx, data_only=True)
    ws = wb["Example"]

    month_09 = {
      "target": ws['D3'].value,
      "actual": ws['E3'].value,
      "miss"  : ws['F3'].value,
      "remain": ws['G3'].value
    }

    month_10 = {
      "target": ws['D4'].value,
      "actual": ws['E4'].value,
      "miss"  : ws['F4'].value,
      "remain": ws['G4'].value
    }

    month_11 = {
      "target": ws['D5'].value,
      "actual": ws['E5'].value,
      "miss"  : ws['F5'].value,
      "remain": ws['G5'].value
    }

    total = {
      "target": ws['D6'].value,
      "actual": ws['E6'].value,
      "miss"  : ws['F6'].value,
      "remain": ws['G6'].value
    }

    short = fullname.replace(self.filepath, '')

    return {
      "timestamp" : time.ctime(),
      "file"      : short,
      "month_09"  : month_09,
      "month_10"  : month_10,
      "month_11"  : month_11,
      "total"     : total
    }

  # Legacy, for use without rich TUI
  def _dump_data_text(self, data):
    print("Timestamp     : %s" % data["timestamp"])
    print("File Modified : %s" % data["file"])
    print("September     : %s" % data["month_09"])
    print("October       : %s" % data["month_10"])
    print("November      : %s" % data["month_11"])
    print("Total         : %s" % data["total"])
    print()

  # Rich TUI related
  def format_table(self, dataTable):
    c = dataTable.columns

    c[1].header_style = "bold blue"
    c[2].header_style = "bold green"
    c[3].header_style = "bold red"  
    c[4].header_style = "bold yellow"  

    c[0].style = "bold"
    c[1].style = "blue"
    c[2].style = "green"
    c[3].style = "red"
    c[4].style = "yellow" 

  def to_str(self, label, m):
    return [label,
      str(m["target"]), str(m["actual"]),
      str(m["miss"]), str(m["remain"])]

  def generate_table(self, data) -> Table:
    # Make a new table
    table = Table(
      "Month", "Target", "Actual", "Miss", "Remain",
      expand=True)

    self.format_table(table)

    table.add_row(*self.to_str(
      "September", data["month_09"]))

    table.add_row(*self.to_str(
      "October",   data["month_10"]))

    table.add_row(*self.to_str(
      "November", data["month_11"]))

    table.add_row(*self.to_str(
      "Total",   data["total"]))

    return table