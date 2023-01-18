import asyncio, openpyxl, time
import xl2web_25

from openpyxl    import load_workbook
from rich.table  import Table

# Excel to Web, Example Class
class Xl2Web_c(xl2web_25.Xl2Web):
  # Excel Reader
  def _pack_data(self, fullname):
    # reopen the worksheet all over again
    wb = load_workbook(self.xlsx, data_only=True)
    ws = wb["Example"]

    month_09 = {
      "budget": ws['E3'].value,
      "actual": ws['E4'].value,
      "gap"   : ws['E5'].value
    }

    month_10 = {
      "budget": ws['G3'].value,
      "actual": ws['G4'].value,
      "gap"   : ws['G5'].value
    }

    short = fullname.replace(self.filepath, '')

    return {
      "timestamp" : time.ctime(),
      "file"      : short,
      "month_09"  : month_09,
      "month_10"  : month_10
    }

  # Legacy, for use without rich TUI
  def _dump_data_text(self, data):
    print("Timestamp     : %s" % data["timestamp"])
    print("File Modified : %s" % data["file"])
    print("September     : %s" % data["month_09"])
    print("October       : %s" % data["month_10"])
    print()

  # Rich TUI related
  def format_table(self, dataTable):
    c = dataTable.columns

    c[1].header_style = "bold blue"
    c[2].header_style = "bold green"
    c[3].header_style = "bold yellow"  

    c[0].style = "bold"
    c[1].style = "blue"
    c[2].style = "green"
    c[3].style = "yellow" 

  def to_str(self, label, m):
    return [label, str(m["budget"]), 
      str(m["actual"]), str(m["gap"])]

  def generate_table(self, data) -> Table:
    # Make a new table
    table = Table(
      "Month", "Budget", "Actual", "Gap",
      expand=True)

    self.format_table(table)

    table.add_row(*self.to_str(
      "September", data["month_09"]))

    table.add_row(*self.to_str(
      "October",   data["month_10"]))

    return table